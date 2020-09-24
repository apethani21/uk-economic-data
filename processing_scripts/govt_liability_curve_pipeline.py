import os
import re
import numpy as np
import pandas as pd
from zipfile import ZipFile
from openpyxl import load_workbook


glc_metadata = {
    "inflation": {
        "path": "../raw_data/glcinflationddata.zip",
        "spreadsheets": ['GLC Inflation daily data_1985 to 1989.xlsx',
                         'GLC Inflation daily data_1990 to 2000.xlsx',
                         'GLC Inflation daily data_1995 to 1999.xlsx',
                         'GLC Inflation daily data_2000 to 2004.xlsx',
                         'GLC Inflation daily data_2005 to 2015.xlsx',
                         'GLC Inflation daily data_2016 to present.xlsx'],
        "sheetname_map": {
            '1. fwds, short end': 'glc_infl_short_end',
            '2. fwd curve': 'glc_infl_forward',
            '3. spot, short end': 'glc_infl_spot_short_end',
            '4. spot curve': 'glc_infl_spot',
            '1.  inf fwds, short end': 'glc_infl_short_end',
            '2.  inf fwd curve': 'glc_infl_forward',
            '3. inf exp spot, short end': 'glc_infl_spot_short_end',
            '4.  inf spot curve': 'glc_infl_spot'
        }
    },
    "nominal": {
        "path": "../raw_data/glcnominalddata.zip",
        "spreadsheets": ['GLC Nominal daily data_1979 to 1984.xlsx',
                         'GLC Nominal daily data_1985 to 1989.xlsx',
                         'GLC Nominal daily data_1990 to 1994.xlsx',
                         'GLC Nominal daily data_1995 to 1999.xlsx',
                         'GLC Nominal daily data_2000 to 2004.xlsx',
                         'GLC Nominal daily data_2005 to 2015.xlsx',
                         'GLC Nominal daily data_2016 to present.xlsx'],
        "sheetname_map": {
            '1. fwds, short end': 'glc_nom_short_end',
            '2. fwd curve': 'glc_nom_forward',
            '3. spot, short end': 'glc_nom_spot_short_end',
            '4. spot curve': 'glc_nom_spot',
            '1. nominal fwds, short end': 'glc_nom_short_end',
            '2. nominal fwd curve': 'glc_nom_forward',
            '3. nominal spot, short end': 'glc_nom_spot_short_end',
            '4. nominal spot curve': 'glc_nom_spot'
        }
    },
    "real": {
        "path": "../raw_data/glcrealddata.zip",
        "spreadsheets": ['GLC Real daily data_1985 to 1989.xlsx',
                         'GLC Real daily data_1990 to 1994.xlsx',
                         'GLC Real daily data_1995 to 1999.xlsx',
                         'GLC Real daily data_2000 to 2004.xlsx',
                         'GLC Real daily data_2005 to 2015.xlsx',
                         'GLC Real daily data_2016 to present.xlsx'],
        "sheetname_map": {
            '1. fwds, short end': 'glc_real_short_end',
            '2. fwd curve': 'glc_real_forward',
            '3. spot, short end': 'glc_real_spot_short_end',
            '4. spot curve': 'glc_real_spot',
            '1. real fwds, short end': 'glc_real_short_end',
            '2. real fwd curve': 'glc_real_forward',
            '2.  real fwd curve': 'glc_real_forward',
            '3. real spot, short end': 'glc_real_spot_short_end',
            '4. real spot curve': 'glc_real_spot',
            '4.  real spot curve': 'glc_real_spot'
        }
    }

}


def _process_spreadsheet(path_to_zip, filename):
    data = {}
    with ZipFile(path_to_zip, "r") as zf:
        with zf.open(filename, "r") as f:
            wb = load_workbook(filename=f, read_only=True, data_only=True)
            sheetnames = [name for name in wb.sheetnames if re.match("\d\.", name)]
            for sheet in sheetnames:
                data[sheet] = [
                    [cell.value for cell in row] for row in wb[sheet].iter_rows(
                        min_row=1, min_col=1, max_row=wb[sheet].max_row, max_col=wb[sheet].max_column
                    )
                ]
    return data


def read_glc_data(data_type):
    data = {}
    path = glc_metadata[data_type]["path"]
    for i, spreadsheet in enumerate(glc_metadata[data_type]["spreadsheets"]):
        data[f"part{i+1}"] = _process_spreadsheet(path, spreadsheet)
    return data


def _process_raw_to_dataframe(data, var):
    dataframes = []
    for part, data_section in data.items():
        df = pd.DataFrame(data_section[var])
        df = df.dropna(axis="columns", how="all")
        df = df.dropna(axis="rows", thresh=df.shape[1]-1)
        df.reset_index(inplace=True, drop=True)
        df.columns = df.iloc[0].values
        if "months:" in df.columns:
            df.rename(columns={"months:": "date"}, inplace=True)
            period_type = "month"
        elif "years:" in df.columns:
            df.rename(columns={"years:": "date"}, inplace=True)
            period_type = "year"
        df = df[2:].set_index("date")
        df.columns = df.columns.map(lambda col: str(round(col, 2)))
        df.columns = df.columns.map(lambda col: col[:-2] if str(col).endswith(".0") else col)
        df.columns = df.columns.map(lambda x: f"{var}_{period_type}_{x}")
        dataframes.append(df)
    return pd.concat(dataframes).fillna(value=np.nan)


def process():
    os.makedirs("../processed_data", exist_ok=True)
    for data_type in glc_metadata:
        sheetname_map = glc_metadata[data_type]["sheetname_map"]
        data = read_glc_data(data_type)
        for part in data.keys():
            for old_name in list(data[part].keys()):
                data[part][sheetname_map[old_name]] = data[part].pop(old_name)
        dataframes = [_process_raw_to_dataframe(data, sheet)
                      for sheet in sheetname_map.values()]
        df = pd.concat(dataframes)
        df.to_parquet(f"../processed_data/glc_{data_type}.parquet")


if __name__ == "__main__":
    process()
    print(os.path.basename(__file__), "done")
