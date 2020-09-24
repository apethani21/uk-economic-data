import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook


def read_hpi_data():
    path_to_file = "../raw_data/UK_House_price_index.xlsx"
    data = {}
    wb = load_workbook(filename=path_to_file, read_only=True, data_only=True)
    sheetnames = [name for name in wb.sheetnames if name != "Metadata"]
    for sheet in sheetnames:
        data[sheet] = [
            [cell.value for cell in row] for row in wb[sheet].iter_rows(
                min_row=1, min_col=1, max_row=wb[sheet].max_row, max_col=wb[sheet].max_column
            )
        ]
    return data


def _clean_col_by_type(col):
    col = "_".join(col.split()).lower()
    col = col.replace("united_kingdom", "uk")
    return col


def _process_raw_to_dataframe_by_type(data):
    df = pd.DataFrame(data['By type'])
    df = df[1:]
    df[0].ffill(inplace=True)
    df = df.transpose()
    df[1].ffill(inplace=True)
    df[2].ffill(inplace=True)
    df = df.transpose()
    df.columns = (df.iloc[:3].replace(np.nan, '').astype(str) + " ").sum(numeric_only=False).apply(str.strip)
    df = df[3:]
    df.dropna(axis="rows", thresh=df.shape[1]-1, inplace=True)
    df["date"] = df.iloc[:, 0].astype(int).astype(str) + "-" + df.iloc[:, 1]
    df["date"] = pd.to_datetime(df["date"])
    df.set_index("date", inplace=True)
    df.drop(columns=[""], inplace=True)
    df.replace("-", np.nan, inplace=True)
    df.columns = df.columns.map(_clean_col_by_type)
    return df


def _process_raw_to_dataframe_non_type(data, sheetname):
    sheetname_map = {
        "Average price": "avg_price",
        "Index Price": "idx_price",
        "Sales Volume": "sales_vol"
    }
    df = pd.DataFrame(data[sheetname])
    df.dropna(axis="columns", how="all", inplace=True)
    df.loc[0, 0] = "date"
    df.columns = df.iloc[0].map(str.lower)
    df.drop([0, 1], inplace=True)
    df["date"] = pd.to_datetime(df["date"])
    df.set_index("date", inplace=True)
    df.dropna(axis="rows", how="all", inplace=True)
    df.columns = df.columns.map(lambda x: x.replace(" ", "_").replace("&", "and"))
    df.columns = df.columns.map(lambda x: f"{x}_{sheetname_map[sheetname]}")
    return df


def process():
    os.makedirs("../processed_data", exist_ok=True)
    data = read_hpi_data()
    dataframes = ([_process_raw_to_dataframe_by_type(data)]
                  + [_process_raw_to_dataframe_non_type(data, sheetname)
                     for sheetname in data.keys() if sheetname != "By type"])
    df = pd.concat(dataframes, axis="columns")
    df.to_parquet("../processed_data/uk_house_price_index.parquet")


if __name__ == "__main__":
    process()
    print(os.path.basename(__file__), "done")
