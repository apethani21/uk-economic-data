import os
import re
import pandas as pd
from zipfile import ZipFile
from openpyxl import load_workbook


def _process_spreadsheet(path_to_zip, filename):
    data = {}
    with ZipFile(path_to_zip, "r") as zf:
        with zf.open(filename, "r") as f:
            wb = load_workbook(filename=f, read_only=True, data_only=True)
            sheetnames = [name for name in wb.sheetnames if re.match("\d\.", name)]
            for sheet in sheetnames:
                data[sheet] = [
                    [cell.value for cell in row] for row in wb[sheet].iter_rows(
                        min_row=1, min_col=1, max_row=wb[sheet].max_row, max_col=wb[sheet].max_column
                    )
                ]
    return data


def read_ois_data():
    path_to_zip = "../raw_data/oisddata.zip"
    data = {}
    data['part1'] = _process_spreadsheet(path_to_zip, "OIS daily data_2009 to 2015.xlsx")
    data['part2'] = _process_spreadsheet(path_to_zip, "OIS daily data_2016 to present.xlsx")
    return data


def _process_raw_to_dataframe(data, var):
    dataframes = []
    for data_section in data.values():
        df = pd.DataFrame(data_section[var])
        df = df.dropna(axis="rows", thresh=df.shape[1]-1)
        df.reset_index(inplace=True, drop=True)
        df.rename(columns={0: "date"}, inplace=True)
        df.columns = df.columns.map(lambda x: f"{var}_month_{x}" if isinstance(x, int) else x)
        df = df[2:].set_index("date")
        dataframes.append(df)
    return pd.concat(dataframes)


def process():
    os.makedirs("../processed_data", exist_ok=True)
    data = read_ois_data()
    sheetname_map = {"1. fwd curve": "ois_forward",
                     "2. spot curve": "ois_spot"}

    for part in data.keys():
        for old_name in list(data[part].keys()):
            data[part][sheetname_map[old_name]] = data[part].pop(old_name)

    df_spot = _process_raw_to_dataframe(data, "ois_spot")
    df_forward = _process_raw_to_dataframe(data, "ois_forward")
    df = df_spot.merge(df_forward, how="outer",
                       left_index=True, right_index=True).astype(float)
    df.to_parquet("../processed_data/ois.parquet")


if __name__ == "__main__":
    process()
    print(os.path.basename(__file__), "done")
