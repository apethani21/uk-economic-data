import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook


def read_short_sterling_pdfs_data():
    path_to_file = "../raw_data/shortsterling_pdfs.xlsx"
    data = {}
    wb = load_workbook(filename=path_to_file, read_only=True, data_only=True)
    sheetnames = [name for name in wb.sheetnames if "month" in name]
    for sheet in sheetnames:
        data[sheet] = [
            [cell.value for cell in row] for row in wb[sheet].iter_rows(
                min_row=1, min_col=1, max_row=wb[sheet].max_row, max_col=wb[sheet].max_column
            )
        ]
    return data


def _clean_col(col, sheetname):
    sheename_map = {'3 month constant maturity': '3mo_mat',
                    '6 month constant maturity': '6mo_mat',
                    '12 month constant maturity': '12mo_mat'}
    prefix = sheename_map.get(sheetname, '')
    main = col.split()[-1].lower()
    if "volatilities" in col:
        return f"{prefix}_short_sterling_implied_vol_{main}"
    elif "LEVEL" in col:
        return f"{prefix}_short_sterling_level_implied_{main}"
    elif "LOGARITHMIC" in col:
        return f"{prefix}_short_sterling_log_implied_{main}"
    elif "PERCENTILES" in col:
        return f"{prefix}_short_sterling_implied_cumprob_{main}"
    elif "Description" in col:
        return "date"
    else:
        raise ValueError(f"Unexpected column {col}")


def _process_raw_to_dataframe(data, sheetname):
    df = pd.DataFrame(data[sheetname]).fillna(value=np.nan)
    df = df.dropna(axis="columns", how="all")
    df = df.dropna(axis="rows", how="all")
    df.drop([0, 1], inplace=True)
    df.reset_index(inplace=True, drop=True)
    df = df.transpose()
    df[0].ffill(inplace=True)
    df = df.transpose()
    df.columns = (df.iloc[:3].replace(np.nan, '').astype(str) + " ").sum(numeric_only=False)
    df.columns = df.columns.map(str.strip)
    df = df[3:]
    df.reset_index(inplace=True, drop=True)
    df.columns = df.columns.map(lambda col: _clean_col(col, sheetname))
    df["date"] = pd.to_datetime(df["date"], yearfirst=True)
    df.set_index("date", inplace=True, drop=True)
    df.replace('', np.nan, inplace=True)
    df = df.dropna(axis="rows", thresh=df.shape[1]-1)
    return df


def process():
    os.makedirs("../processed_data", exist_ok=True)
    data = read_short_sterling_pdfs_data()
    df_3mo = _process_raw_to_dataframe(data, '3 month constant maturity')
    df_6mo = _process_raw_to_dataframe(data, '6 month constant maturity')
    df_12mo = _process_raw_to_dataframe(data, '12 month constant maturity')
    df = (df_3mo
          .merge(df_6mo, how="outer", left_index=True, right_index=True)
          .merge(df_12mo, how="outer", left_index=True, right_index=True)).astype(float)
    df.to_parquet("../processed_data/short_sterling_pdfs.parquet")


if __name__ == "__main__":
    process()
    print(os.path.basename(__file__), "done")
