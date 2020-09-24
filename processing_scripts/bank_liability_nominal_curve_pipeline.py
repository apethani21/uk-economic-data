import os
import re
import numpy as np
import pandas as pd
from zipfile import ZipFile
from openpyxl import load_workbook


def _process_spreadsheet(path_to_zip, filename):
    data = {}
    with ZipFile(path_to_zip, "r") as zf:
        with zf.open(filename, "r") as f:
            wb = load_workbook(filename=f, read_only=True, data_only=True)
            sheetnames = [name for name in wb.sheetnames if re.match("\d\.", name)]
            for sheet in sheetnames:
                data[sheet] = [
                    [cell.value for cell in row] for row in wb[sheet].iter_rows(
                        min_row=1, min_col=1, max_row=wb[sheet].max_row, max_col=wb[sheet].max_column
                    )
                ]
    return data


def read_glc_real_data():
    path_to_zip = "../raw_data/blcnomddata.zip"
    data = {}
    spreadsheets = ['BLC Nominal daily data_1990 to 1994.xlsx',
                    'BLC Nominal daily data_1995 to 1999.xlsx',
                    'BLC Nominal daily data_2000 to 2004.xlsx',
                    'BLC Nominal daily data_2005 to 2015.xlsx',
                    'BLC Nominal daily data_2016 to present.xlsx']
    for i, spreadsheet in enumerate(spreadsheets):
        data[f"part{i+1}"] = _process_spreadsheet(path_to_zip, spreadsheet)
    return data


def _process_raw_to_dataframe(data, var):
    dataframes = []
    for part, data_section in data.items():
        df = pd.DataFrame(data_section[var])
        df = df.dropna(axis="columns", how="all")
        df = df.dropna(axis="rows", thresh=df.shape[1]-1)
        df.reset_index(inplace=True, drop=True)
        df.columns = df.iloc[0].values
        if "months:" in df.columns:
            df.rename(columns={"months:": "date"}, inplace=True)
            period_type = "month"
        elif "years:" in df.columns:
            df.rename(columns={"years:": "date"}, inplace=True)
            period_type = "year"
        df = df[2:].set_index("date")
        df.columns = df.columns.map(lambda col: str(round(col, 2)))
        df.columns = df.columns.map(lambda col: col[:-2] if str(col).endswith(".0") else col)
        df.columns = df.columns.map(lambda x: f"{var}_{period_type}_{x}")
        dataframes.append(df)
    return pd.concat(dataframes).fillna(value=np.nan)


def process():
    os.makedirs("../processed_data", exist_ok=True)
    data = read_glc_real_data()
    sheetname_map = {
        '1. fwds, short end': 'blc_nom_short_end',
        '2. fwd curve': 'blc_nom_forward',
        '3. spot, short end': 'blc_nom_spot_short_end',
        '4. spot curve': 'blc_nom_spot',
        '1. BLC fwds, short end': 'blc_nom_short_end',
        '2. BLC fwd curve': 'blc_nom_forward',
        '3. BLC spot, short end': 'blc_nom_spot_short_end',
        '4. BLC spot curve': 'blc_nom_spot'
    }

    for key in data.keys():
        for old_name in list(data[key].keys()):
            data[key][sheetname_map[old_name]] = data[key].pop(old_name)
    dataframes = [_process_raw_to_dataframe(data, sheet) for sheet in sheetname_map.values()]
    df = pd.concat(dataframes)
    df.to_parquet("../processed_data/bank_liability_curve_nominal.parquet")


if __name__ == "__main__":
    process()
    print(os.path.basename(__file__), "done")
